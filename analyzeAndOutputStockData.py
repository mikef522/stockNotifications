import datetime
import yfinance as yf

import pandas_datareader as pdr

import pandas as pd
import smtplib

import schedule
import time

from datetime import datetime
from datetime import timedelta
from pytz import timezone

import math


from stockDataRetrieval import getStockData


import numpy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from graphingStockCode import graphStockDataAndReturnStockDataUpdatedWithIntersectionData


def getQuarterandYearNameFromQuarterEndDate(d):
    if (d < datetime(d.year,3,31)):#years are the same, so I'm only comparing month and day
        return 'Q4 ' + str(d.year-1)
    elif (d < datetime(d.year,6,30)):
        return 'Q1 ' + str(d.year)
    elif (d < datetime(d.year,9,30)):
        return 'Q2 ' + str(d.year)
    elif (d < datetime(d.year,12,31)):
        return 'Q3 ' + str(d.year)
    return 'Q4 ' + str(d.year)






def analyzeAndOutputStockData(stockSymbolList,dateOfInterest, userOptions):
    getStockDataOutput = getStockData(stockSymbolList, dateOfInterest)

    listOfStockData = getStockDataOutput[0]

    '''
    ahistory = yahooQueryStock.history(start=date - timedelta(days=365 * 2), end=date + timedelta(days=1))
    graphStockData(ahistory, stock.stockSymbol, stock.desiredDateForStockData)
    '''

    graphPdfFileName, listOfStockData = graphStockDataAndReturnStockDataUpdatedWithIntersectionData(listOfStockData, listOfStockData[0].desiredDateForStockData, userOptions)#default graph period is 1 year
    #print(listOfStockData[0].intersectionData)


    #sorting data and putting nan sorting value data on bottom. could sort the nan data by something else if I really wanted.
    listOfStockDataWithNanSortingValue = [ x for x in listOfStockData if (math.isnan(x.sortingValue) == True) ]
    listOfStockDataWithValidSortingValue = [ x for x in listOfStockData if (math.isnan(x.sortingValue) == False) ]
    listOfStockDataWithValidSortingValue.sort(key=lambda x: x.sortingValue, reverse=True)
    listOfStockDataWithValidSortingValue.extend(listOfStockDataWithNanSortingValue)
    completelySortedListOfStockData = listOfStockDataWithValidSortingValue

    listOfAllQuarterDates = getStockDataOutput[1]
    listOfAllQuarterDatesSorted = sorted(listOfAllQuarterDates, reverse=True)









    #create pandas data frame for all the stock data
    stockData2DListForDataFrame = []

    ##
    ##COLUMN HEADERS
    initialColumnTitles = ['Max ago change %','10yr(uses 1mo avg) change %','5yr(uses 1mo avg) change %','March low change(%)', 'Stock', 'Price', 'Change($)', 'Change(%)', 'PE Ratio']
    stockDataColumnTitlesForDataFrame = initialColumnTitles.copy()

    sortedListOfAllQuarterNames = []

    #listOfAllQuarterDatesSorted = (datetime(2020,12,31),datetime(2020,9,30),datetime(2020,6,30),datetime(2020,3,31),datetime(2019,12,31)) #for testing

    #sort all of the quarterly report dates into a list of appropriate quarter and year names. This list will be sorted since the list of quarter dates is sorted
    for i in range(len(listOfAllQuarterDatesSorted)):
        quarterAndYearName = getQuarterandYearNameFromQuarterEndDate(listOfAllQuarterDatesSorted[i])
        if ((quarterAndYearName in sortedListOfAllQuarterNames) == False):
            sortedListOfAllQuarterNames.append(quarterAndYearName)

    quarterlyDataColumnTitles = []

    for i in range(len(sortedListOfAllQuarterNames)):
        quarterlyDataColumnTitles.append(sortedListOfAllQuarterNames[i] + ' Revenue')
        quarterlyDataColumnTitles.append(sortedListOfAllQuarterNames[i] + ' Profits')
        quarterlyDataColumnTitles.append(sortedListOfAllQuarterNames[i] + ' Margins')

    stockDataColumnTitlesForDataFrame.extend(quarterlyDataColumnTitles)

    currentTimeAndDate = datetime.now(timezone('US/Eastern'))
    additionalColumnTitles = ['Stock2', 'Financial Data Currency', 'Conversion factors that were used to convert financial data to USD (left to right, latest to oldest quarter)',
                              currentTimeAndDate.strftime('Latest Market cap (as of %m-%d-%Y)'), currentTimeAndDate.strftime('Latest Volume (as of %m-%d-%Y)'),#these 2 must stay together in this order
                              currentTimeAndDate.strftime('Latest Dividend (as of %m-%d-%Y)'), currentTimeAndDate.strftime('Latest Dividend Yield (as of %m-%d-%Y)'),
                              'last date 50MA up over 200MA', 'last date 50MA down under 200MA',
                              'last date 14MA up over 200MA', 'last date 14MA down under 200MA',
                              'last date price up over 200MA', 'last date price down under 200MA',
                              ]#these 2 must stay together in this order, Stock2 used because pandas dataframe gets confused with duplicate column names and will duplicate columns
    stockDataColumnTitlesForDataFrame.extend(additionalColumnTitles)


    ##
    ##ORGANIZE STOCK DATA ROW BY ROW
    for stock in completelySortedListOfStockData:
        rFac = 1000#reduction factor for reducing large numbers like revenue and income

        stockDataRowAsList = []

        ##
        ##ADDING PRE-QUARTERLY DATA

        if (math.isnan(stock.closePriceMaxAgo) == True):
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append((stock.closePrice - stock.closePriceMaxAgo)/stock.closePriceMaxAgo)
        if (math.isnan(stock.meanPriceForMonthTenYearsAgo) == True):
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append((stock.closePrice - stock.meanPriceForMonthTenYearsAgo)/stock.meanPriceForMonthTenYearsAgo)
        if(math.isnan(stock.meanPriceForMonthFiveYearsAgo) == True):
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append((stock.closePrice - stock.meanPriceForMonthFiveYearsAgo)/stock.meanPriceForMonthFiveYearsAgo)


        if (math.isnan(stock.sortingValue) == True):#if no data from march, write N/A
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append(stock.sortingValue*100)

        stockDataRowAsList.extend([stock.stockSymbol, stock.closePrice, stock.dayPriceChangeDollars, stock.priceChangePercent*100])

        if (math.isnan(stock.peRatio) == True):#if no pe ratio, write N/A
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append(stock.peRatio)

        ##
        ##ADDING QUARTERLY DATA
        quarterlyData = []
        '''
        for i in range(len(listOfAllQuarterDatesSorted)):
            quarterlyData.append('No Data')

        matchingDateIndexesSorted = [key for key, val in enumerate(listOfAllQuarterDatesSorted) if val in set(stock.quarterDateUpToFourQ)]
        for i in range(len(matchingDateIndexesSorted)):
            quarterlyData[matchingDateIndexesSorted[i]] = ['${:<15,.0f}'.format(stock.revenueUpToFourQ[i]/rFac), '${:<15,.0f}'.format(stock.netProfitUpToFourQ[i]/rFac), '{:<+7.2f}%'.format(stock.netMarginUpToFourQ[i]*100)]
        '''
        #initialize quarterly data in row to 'No Data'
        for i in range(len(quarterlyDataColumnTitles)):
            quarterlyData.append('No Data')


        for i in range(len(sortedListOfAllQuarterNames)):
            #check to see if current quarter name exists in data for this stock. if it does, fill in quarterly data for that quarter
            for j in range(len(stock.quarterDateUpToFourQ)):
                if (getQuarterandYearNameFromQuarterEndDate(stock.quarterDateUpToFourQ[j]) == sortedListOfAllQuarterNames[i]):
                    #length of sortedListOfAllQuarterNames is 1/3 that of quarterlyDataColumnTitles, so account for that difference in filling in revenue, profit, and margin data
                    #perform currency conversion as necessary
                    if ((stock.financialCurrency != 'USD') and (stock.financialCurrency != '') and (len(stock.conversionFactorToConvertToUSD) != 0)):
                            quarterlyData[i*3] = (stock.revenueUpToFourQ[j]*stock.conversionFactorToConvertToUSD[j]) / rFac
                            quarterlyData[(i*3)+1] = (stock.netProfitUpToFourQ[j]*stock.conversionFactorToConvertToUSD[j]) / rFac

                    else:
                        quarterlyData[i * 3] = stock.revenueUpToFourQ[j] / rFac
                        quarterlyData[(i * 3) + 1] = stock.netProfitUpToFourQ[j] / rFac

                    quarterlyData[(i * 3) + 2] = stock.netMarginUpToFourQ[j] * 100


        stockDataRowAsList.extend(quarterlyData)



        ##
        ##ADDING ADDITIONAL DATA
        stockDataRowAsList.append(stock.stockSymbol)
        if (stock.financialCurrency == ''):
            stockDataRowAsList.append('N/A')
            stockDataRowAsList.append('N/A')
        else:
            if (stock.financialCurrency == 'USD'):
                stockDataRowAsList.append(stock.financialCurrency)
            elif (len(stock.conversionFactorToConvertToUSD) == 0):
                stockDataRowAsList.append(stock.financialCurrency + ' NO CONVERSION AVAILABLE, NOT CONVERTED')
            else:
                stockDataRowAsList.append(stock.financialCurrency + ', now converted to USD')
            stockDataRowAsList.append(list(numpy.around(numpy.array(stock.conversionFactorToConvertToUSD), 4)))#round conversion factor to 4 decimal places

        if (math.isnan(stock.marketCap)):
            stockDataRowAsList.append('No data')
        else:
            stockDataRowAsList.append(stock.marketCap)

        if (math.isnan(stock.volume)):
            stockDataRowAsList.append('No data')
        else:
            stockDataRowAsList.append(stock.volume)

        if (math.isnan(stock.dividendRate)):
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append(stock.dividendRate)

        if (math.isnan(stock.dividendYield)):
            stockDataRowAsList.append('N/A')
        else:
            stockDataRowAsList.append(stock.dividendYield)

        noLastDateMessage = 'None within past year'
        if (len(stock.intersectionData['50dayMAOver200dayMA']['upCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['50dayMAOver200dayMA']['upCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)

        if (len(stock.intersectionData['50dayMAOver200dayMA']['downCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['50dayMAOver200dayMA']['downCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)

        if (len(stock.intersectionData['14dayMAOver200dayMA']['upCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['14dayMAOver200dayMA']['upCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)

        if (len(stock.intersectionData['14dayMAOver200dayMA']['downCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['14dayMAOver200dayMA']['downCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)

        if (len(stock.intersectionData['priceOver200dayMA']['upCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['priceOver200dayMA']['upCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)

        if (len(stock.intersectionData['priceOver200dayMA']['downCrossDates']) > 0):
            stockDataRowAsList.append(stock.intersectionData['priceOver200dayMA']['downCrossDates'][-1])
        else:
            stockDataRowAsList.append(noLastDateMessage)






        ##
        ##ADD DATA FOR CURRENT ROW
        stockData2DListForDataFrame.append(stockDataRowAsList)




    ##
    ##SAVE EXCEL FILE

    #create data frame representing all stock data gathered
    stockData_DataFrame = pd.DataFrame(stockData2DListForDataFrame, columns=stockDataColumnTitlesForDataFrame)



    # if I want revenues, profits and margins all next to eachother, change the column order
    cols_stockData_DataFrameRef = list(stockData_DataFrame.columns.values)

    newQuarterlyColOrder = []
    newQuarterlyColOrder.extend(quarterlyDataColumnTitles[0::3])#revenues
    newQuarterlyColOrder.extend(quarterlyDataColumnTitles[1::3])#margins
    newQuarterlyColOrder.extend(quarterlyDataColumnTitles[2::3])#profits

    #numberOfColumnsBeforeQuarterlyCols = len(initialColumnTitles)
    newColsOrder = cols_stockData_DataFrameRef[:len(initialColumnTitles)]
    newColsOrder.extend(newQuarterlyColOrder)
    newColsOrderLength = len(newColsOrder)
    newColsOrder.extend(cols_stockData_DataFrameRef[newColsOrderLength:len(additionalColumnTitles)+newColsOrderLength])


    stockData_DataFrame = stockData_DataFrame[newColsOrder]#if you have duplicate column names, that column will be duplicated


    stockDataExcelFileName = 'No excel file saved'
    if userOptions['saveExcel']:
        #save excel file
        currentTimeAndDate = datetime.now(timezone('US/Eastern'))
        stockDataExcelFileName = 'stockData_'+dateOfInterest.strftime('%m-%d-%Y___') + currentTimeAndDate.strftime('generated_%m-%d-%Y_%I-%M%p_EST') + '.xlsx'
        stockData_DataFrame.to_excel(stockDataExcelFileName, index=False)




        ##
        ##FORMAT EXCEL FILE using Openpyxl

        workbook = openpyxl.load_workbook(filename=stockDataExcelFileName)
        sht1 = workbook['Sheet1']

        #format data columns before quarterly data
        #sht1.column_dimensions['A'].number_format = '[Color10]+0.00\%;[Red]-0.0#\%'#sorting value, change % from march lows in this case
        ['Max ago change %', '10yr(uses 1mo avg) change %', '5yr(uses 1mo avg) change %', 'March low change(%)', 'Stock',
         'Price', 'Change($)', 'Change(%)', 'PE Ratio']

        ##
        ##Replace this set of code with a function
        maxAgoChangePercColLetter = get_column_letter(1 + initialColumnTitles.index('Max ago change %'))
        for col_cell in sht1[maxAgoChangePercColLetter]:
            col_cell.number_format = '[Color10]+0.00%;[Red]-0.00%'
        tenYearChangePercColLetter = get_column_letter(1 + initialColumnTitles.index('10yr(uses 1mo avg) change %'))
        for col_cell in sht1[tenYearChangePercColLetter]:
            col_cell.number_format = '[Color10]+0.00%;[Red]-0.00%'
        fiveYearChangePercColLetter = get_column_letter(1 + initialColumnTitles.index('5yr(uses 1mo avg) change %'))
        for col_cell in sht1[fiveYearChangePercColLetter]:
            col_cell.number_format = '[Color10]+0.00%;[Red]-0.00%'

        sortingValueColLetter = get_column_letter(1+initialColumnTitles.index('March low change(%)'))
        for col_cell in sht1[sortingValueColLetter]:
            col_cell.number_format = '[Color10]+0.00\%;[Red]-0.00\%'#sorting value, change % from march lows in this case
        priceColLetter = get_column_letter(1 + initialColumnTitles.index('Price'))
        for col_cell in sht1[priceColLetter]:
            col_cell.number_format = '$0.00;-$0.00'  # stock price
        changeDollarColLetter = get_column_letter(1 + initialColumnTitles.index('Change($)'))
        for col_cell in sht1[changeDollarColLetter]:
            col_cell.number_format = '[Color10]+$0.00;[Red]-$0.00'  # change in dollars
        changePercColLetter = get_column_letter(1 + initialColumnTitles.index('Change(%)'))
        for col_cell in sht1[changePercColLetter]:
            col_cell.number_format = '[Color10]+0.00\%;[Red]-0.00\%'  # change in %
        peRatioColLetter = get_column_letter(1 + initialColumnTitles.index('PE Ratio'))
        for col_cell in sht1[peRatioColLetter]:
            col_cell.number_format = '0.00;-0.00'  # pe ratio

        # this formats quarterly data if it's in the all 4 Q rev, all 4 Q profit, all 4 Q margin order
        numberOfColumnsBeforeQuarterlyData = len(initialColumnTitles)
        for j in range(len(quarterlyDataColumnTitles)):
            if (j < len(sortedListOfAllQuarterNames) * 2):
                for col_cell in sht1[get_column_letter(j + 1 + numberOfColumnsBeforeQuarterlyData)]:
                    col_cell.number_format = '[Color10]0,00;[Red]-0,00'

            else:
                for col_cell in sht1[get_column_letter(j + 1 + numberOfColumnsBeforeQuarterlyData)]:
                    col_cell.number_format = '[Color10]0.00\%;[Red]-0.00\%'

        # formatting additional data columns
        numberOfcolumnsBeforeAdditionalData = len(initialColumnTitles)+len(quarterlyDataColumnTitles)

        numberColBeforeMarketCapInAddtlData = additionalColumnTitles.index(currentTimeAndDate.strftime('Latest Market cap (as of %m-%d-%Y)'))
        marketCapAndVolNumberFormat = '0,00;-0,00'
        for col_cell in sht1[get_column_letter(1+numberOfcolumnsBeforeAdditionalData+numberColBeforeMarketCapInAddtlData)]:
            col_cell.number_format = marketCapAndVolNumberFormat
        for col_cell in sht1[get_column_letter(1+numberOfcolumnsBeforeAdditionalData + numberColBeforeMarketCapInAddtlData+1)]:
            col_cell.number_format = marketCapAndVolNumberFormat

        numberColBeforeDividendInAddtlData = additionalColumnTitles.index(currentTimeAndDate.strftime('Latest Dividend (as of %m-%d-%Y)'))
        for col_cell in sht1[get_column_letter(1+numberOfcolumnsBeforeAdditionalData+numberColBeforeDividendInAddtlData)]:
            col_cell.number_format = '$0.00;-$0.00'#dividend format
        for col_cell in sht1[get_column_letter(1+numberOfcolumnsBeforeAdditionalData+numberColBeforeDividendInAddtlData+1)]:
            col_cell.number_format = '0.00%;-0.00%'







        from openpyxl.styles import Font, Border, Side
        #adding additional stock symbol columns
        #insert stock symbol columns in quarterly data
        stockSymbolColNum = 1 + initialColumnTitles.index('Stock')####Note that this will be invalid if I add a stock row before the first row


        ####Cannot use col number before first stock column on left side, otherwise will mess data up
        quarterlyDataNumColToStep = int(len(quarterlyDataColumnTitles)/3)
        firstRevenueColNum = 1+len(initialColumnTitles)
        firstNetIncomeColNum = 1 + len(initialColumnTitles) + quarterlyDataNumColToStep + 1##last 1 counts for inserted stock Row
        firstMarginColNum = 1 + len(initialColumnTitles) + (quarterlyDataNumColToStep*2) + 2#last 2 counts for inserted stock Rows
        stockColInsertionLocationsAfterFirstStockColOnly_LefttoRight = [firstRevenueColNum,firstNetIncomeColNum,firstMarginColNum]

        for colNum in stockColInsertionLocationsAfterFirstStockColOnly_LefttoRight:
            sht1.insert_cols(colNum)
            stockTitleCell = sht1.cell(row=1, column=colNum)
            stockTitleCell.value = 'Stock'
            stockTitleCell.font = Font(bold=True)
            stockTitleCell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for i in range(2, len(stockSymbolList) + 2):
                sht1.cell(row=i, column=colNum).value = sht1.cell(row=i, column=stockSymbolColNum).value

        numStockColsAdded = len(stockColInsertionLocationsAfterFirstStockColOnly_LefttoRight)




        #set alignment of data
        for i in range(1, sht1.max_column + 1):#aligning all data
            for col_cell in sht1[get_column_letter(i)]:
                col_cell.alignment = Alignment(horizontal='right')

        stockSymbolColLetter = get_column_letter(1 + initialColumnTitles.index('Stock'))
        for col_cell in sht1[stockSymbolColLetter]:
            col_cell.alignment = Alignment(horizontal='left')

        numberColBeforeFinCurConvInAddtlData = additionalColumnTitles.index('Conversion factors that were used to convert financial data to USD (left to right, latest to oldest quarter)')
        for col_cell in sht1[get_column_letter(1+numberOfcolumnsBeforeAdditionalData + numStockColsAdded + numberColBeforeFinCurConvInAddtlData)]:
            col_cell.alignment = Alignment(horizontal='left')

        '''
        for i in range(1, sht1.max_column + 1):
            sht1.cell(row=1, column=i).alignment = Alignment(horizontal='center')
    
        # adjusting text wrapping and column width/height
        for i in range((numberOfcolumnsBeforeAdditionalData + 1), (numberOfcolumnsBeforeAdditionalData + 6)):
            sht1.cell(row=1, column=i).alignment = Alignment(wrapText=True)
        '''

        #adjusts column width
        dims = {}
        for row in sht1.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sht1.column_dimensions[col].width = value + 1.23 #1.23 is fudge factor




        workbook.save(filename=stockDataExcelFileName)


        print('\n\n-------\n' + dateOfInterest.strftime('%m-%d-%Y') + ' stock report saved excel file' + '\n-------\n\n\n\n')

    print('\n\n-------\n' + dateOfInterest.strftime('%m-%d-%Y') + ' stock report finished' + '\n-------\n\n\n\n')

    return stockDataExcelFileName, graphPdfFileName

